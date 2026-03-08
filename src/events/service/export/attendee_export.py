"""Generate Excel export of event attendee list."""

import typing as t
from io import BytesIO
from uuid import UUID

import structlog
from openpyxl import Workbook

from common.models import FileExport
from common.service.export_service import complete_export, fail_export, start_export
from events.models import Event, EventRSVP, Ticket

from .formatting import auto_fit_columns, compute_pronoun_distribution, style_header_row, style_summary_sheet

logger = structlog.get_logger(__name__)


def generate_attendee_export(export_id: UUID) -> None:
    """Build an Excel workbook with attendee data for an event."""
    export = FileExport.objects.select_related("requested_by").get(pk=export_id)
    start_export(export)
    try:
        file_bytes = _build_attendee_workbook(export)
        complete_export(export, file_bytes, f"attendee_export_{export_id}.xlsx")
        logger.info("attendee_export_completed", export_id=str(export_id))
    except Exception as e:
        fail_export(export, f"Export failed: {e}")
        logger.exception("attendee_export_failed", export_id=str(export_id))
        raise


def _build_attendee_workbook(export: FileExport) -> bytes:
    """Build the Excel workbook and return raw bytes."""
    event_id = UUID(export.parameters["event_id"])
    event = Event.objects.select_related("venue", "organization").get(pk=event_id)

    tickets = list(
        Ticket.objects.filter(
            event=event,
            status__in=[Ticket.TicketStatus.ACTIVE, Ticket.TicketStatus.CHECKED_IN],
        ).select_related("user", "tier", "seat", "seat__sector", "payment")
    )

    rsvps = list(EventRSVP.objects.filter(event=event, status=EventRSVP.RsvpStatus.YES).select_related("user"))

    ticket_checked_in = sum(1 for tk in tickets if tk.status == Ticket.TicketStatus.CHECKED_IN)

    wb = Workbook()
    _write_summary_sheet(wb, event, tickets, rsvps, ticket_checked_in)
    _write_attendees_sheet(wb, tickets, rsvps)

    buf = BytesIO()
    wb.save(buf)
    file_bytes = buf.getvalue()
    buf.close()
    wb.close()
    return file_bytes


def _write_summary_sheet(
    wb: Workbook,
    event: Event,
    tickets: list[Ticket],
    rsvps: list[EventRSVP],
    ticket_checked_in: int,
) -> None:
    """Populate the Summary sheet."""
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = "Summary"

    # Pronoun distribution (deduplicated by user)
    def _user_pairs() -> t.Iterator[tuple[UUID, t.Any]]:
        for ticket in tickets:
            if ticket.user:
                yield ticket.user_id, ticket.user
        for rsvp in rsvps:
            if rsvp.user:
                yield rsvp.user_id, rsvp.user

    pronoun_stats = compute_pronoun_distribution(_user_pairs())
    sorted_pronouns = pronoun_stats.sorted_pronouns
    total_with_pronouns = pronoun_stats.total_with
    total_without_pronouns = pronoun_stats.total_without

    summary_rows: list[tuple[str, t.Any]] = [
        ("Event", event.name),
        ("Date", event.start.isoformat() if event.start else "N/A"),
        ("Venue", event.venue.name if event.venue else "N/A"),
        ("Organization", event.organization.name),
        ("Total attendees", len(tickets) + len(rsvps)),
        ("Tickets", len(tickets)),
        ("RSVPs", len(rsvps)),
        ("Checked in", ticket_checked_in),
        ("", ""),
        ("Pronoun Distribution", ""),
        ("Total with pronouns", total_with_pronouns),
        ("Total without pronouns", total_without_pronouns),
    ]
    for row in summary_rows:
        ws.append(row)
    for pronouns, count in sorted_pronouns:
        ws.append((f"  {pronouns}", count))

    style_summary_sheet(ws)


_STATUS_DISPLAY: dict[str, str] = {
    Ticket.TicketStatus.ACTIVE: "Active",
    Ticket.TicketStatus.CHECKED_IN: "Checked In",
    Ticket.TicketStatus.CANCELLED: "Cancelled",
    Ticket.TicketStatus.PENDING: "Pending",
}

_RSVP_STATUS_DISPLAY: dict[str, str] = {
    EventRSVP.RsvpStatus.YES: "Yes",
    EventRSVP.RsvpStatus.NO: "No",
    EventRSVP.RsvpStatus.MAYBE: "Maybe",
}


def _write_attendees_sheet(wb: Workbook, tickets: list[Ticket], rsvps: list[EventRSVP]) -> None:
    """Populate the Attendees sheet."""
    ws = wb.create_sheet("Attendees")
    headers = [
        "Name",
        "Email",
        "Pronouns",
        "Type",
        "RSVP Status",
        "Ticket Tier",
        "Ticket Status",
        "Checked In",
        "Checked In At",
        "Guest Name",
        "Seat",
        "Payment",
    ]
    ws.append(headers)

    for ticket in tickets:
        payment = getattr(ticket, "payment", None)
        seat_label = ticket.seat.label if ticket.seat else ""
        is_checked_in = ticket.status == Ticket.TicketStatus.CHECKED_IN
        ws.append(
            [
                ticket.user.get_full_name() if ticket.user else "",
                ticket.user.email if ticket.user else "",
                ticket.user.pronouns if ticket.user else "",
                "Ticket",
                "",
                ticket.tier.name if ticket.tier else "",
                _STATUS_DISPLAY.get(ticket.status, ticket.status),
                "Yes" if is_checked_in else "No",
                ticket.checked_in_at.isoformat() if ticket.checked_in_at else "",
                ticket.guest_name or "",
                seat_label,
                payment.status if payment else (ticket.tier.payment_method if ticket.tier else ""),
            ]
        )

    for rsvp in rsvps:
        ws.append(
            [
                rsvp.user.get_full_name() if rsvp.user else "",
                rsvp.user.email if rsvp.user else "",
                rsvp.user.pronouns if rsvp.user else "",
                "RSVP",
                _RSVP_STATUS_DISPLAY.get(rsvp.status, rsvp.status) if rsvp.status else "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )

    style_header_row(ws)
    auto_fit_columns(ws)
