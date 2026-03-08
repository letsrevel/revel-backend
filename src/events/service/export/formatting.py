"""Shared Excel formatting utilities for export services."""

import typing as t
from uuid import UUID

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

if t.TYPE_CHECKING:
    from accounts.models import RevelUser


class PronounStats(t.NamedTuple):
    """Result of computing pronoun distribution.

    Attributes:
        sorted_pronouns: List of (pronoun, count) tuples sorted by count descending.
        total_with: Total users who have specified pronouns.
        total_without: Total users without pronouns specified.
    """

    sorted_pronouns: list[tuple[str, int]]
    total_with: int
    total_without: int


def compute_pronoun_distribution(users: t.Iterable[tuple[UUID, "RevelUser"]]) -> PronounStats:
    """Compute pronoun distribution from an iterable of (user_id, user) pairs.

    Deduplicates by user_id. Users should already be resolved (not None).
    """
    pronoun_counts: dict[str, int] = {}
    seen_user_ids: set[UUID] = set()
    for user_id, user in users:
        if user_id not in seen_user_ids:
            seen_user_ids.add(user_id)
            key = user.pronouns or ""
            pronoun_counts[key] = pronoun_counts.get(key, 0) + 1

    total_without = pronoun_counts.pop("", 0)
    total_with = sum(pronoun_counts.values())
    sorted_pronouns = sorted(pronoun_counts.items(), key=lambda x: x[1], reverse=True)
    return PronounStats(sorted_pronouns, total_with, total_without)


HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)

LABEL_FONT = Font(bold=True, size=11)

MIN_WIDTH = 10
MAX_WIDTH = 60


def style_header_row(ws: Worksheet) -> None:
    """Apply bold white-on-blue styling to the first row."""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = ws.dimensions


def auto_fit_columns(ws: Worksheet) -> None:
    """Set column widths based on content, clamped to [MIN_WIDTH, MAX_WIDTH]."""
    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), 1):
        max_len = 0
        for cell in col_cells:
            if cell.value is not None:
                # For wrapped text, use the longest line
                text = str(cell.value)
                line_len = max((len(line) for line in text.split("\n")), default=0)
                max_len = max(max_len, line_len)
        width = min(max(max_len + 3, MIN_WIDTH), MAX_WIDTH)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def style_summary_sheet(ws: Worksheet) -> None:
    """Style a summary sheet: bold labels in column A, fixed widths for label/value columns."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        cell = row[0]
        if cell.value:
            cell.font = LABEL_FONT
    # Set reasonable widths for label/value columns
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
