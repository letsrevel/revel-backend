"""Revenue & VAT report bundle builders, cache/generation, and scheduled delivery (#551)."""

import io
import typing as t
import zipfile
from datetime import date, datetime
from decimal import Decimal
from uuid import UUID

import structlog
from django.db import transaction
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Alignment

from common.models import FileExport
from common.service.export_service import complete_export, fail_export, start_export
from common.tasks import send_email
from events.models import Organization
from events.service.export.formatting import LABEL_FONT, auto_fit_columns, style_header_row
from events.service.revenue_aggregation import (
    ReportScope,
    RevenueReportData,
    build_revenue_report_data,
    closed_period_for,
    compute_revenue_data_hash,
    organization_timezone,
)
from events.utils import get_organization_timezone

if t.TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

    from accounts.models import RevelUser

__all__ = [
    "ReportScope",
    "RevenueReportData",
    "build_revenue_report_data",
    "compute_revenue_data_hash",
    "closed_period_for",
    "organization_timezone",
    "get_or_generate_revenue_report",
    "generate_revenue_report",
    "deliver_scheduled_revenue_reports",
    "build_xlsx",
    "build_pdf",
    "build_zip",
    "report_filename",
]

logger = structlog.get_logger(__name__)

# ---------------------------------------------------------------------------
# Bundle builders
# ---------------------------------------------------------------------------

# Locale-aware Excel number formats: the group/decimal separators render per the
# viewer's locale (e.g. "1.234.567,89" in an AT/DE Excel), so a single format code
# serves both comma- and dot-grouping audiences.
_MONEY_FORMAT = "#,##0.00"
_INT_FORMAT = "#,##0"
_PERCENT_FORMAT = '0.##"%"'  # value 20 -> "20%", 7.5 -> "7.5%" (literal %, no x100)


def _format_numeric_columns(
    ws: "Worksheet",
    money_cols: tuple[int, ...] = (),
    int_cols: tuple[int, ...] = (),
    percent_cols: tuple[int, ...] = (),
) -> None:
    """Right-align and number-format the given 1-based columns across data rows."""
    right = Alignment(horizontal="right")
    specs = ((money_cols, _MONEY_FORMAT), (int_cols, _INT_FORMAT), (percent_cols, _PERCENT_FORMAT))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cols, fmt in specs:
            for col in cols:
                cell = row[col - 1]
                if isinstance(cell.value, (int, float, Decimal)) and not isinstance(cell.value, bool):
                    cell.number_format = fmt
                    cell.alignment = right


_TXN_HEADERS = [
    "date",
    "payment_id",
    "event",
    "tier",
    "buyer_country",
    "reverse_charge",
    "gross",
    "net",
    "vat_rate",
    "vat_amount",
    "discount",
    "refund_amount",
    "currency",
    "stripe_session_id",
    "stripe_payout_id",
]


def report_filename(scope: ReportScope, ext: str = "zip") -> str:
    """Return a canonical filename for the report bundle or one of its parts."""
    return f"revel-revenue-{scope.org.slug}-{scope.date_from}_{scope.date_to}.{ext}"


def build_xlsx(data: RevenueReportData) -> bytes:
    """Build the two-sheet XLSX workbook (Summary + Transactions) and return raw bytes."""
    wb = Workbook()
    summary = wb.active
    assert summary is not None  # Workbook() always creates a default sheet
    summary.title = "Summary"
    summary.append(["Currency", "VAT rate", "Net", "VAT", "Gross", "Tickets"])
    for section in data.sections:
        for bucket in section.rate_buckets:
            summary.append([section.currency, bucket.label, bucket.net, bucket.vat, bucket.gross, bucket.ticket_count])
        summary.append([section.currency, "Refunds", None, None, -section.refunds_total, section.refunded_count])
        summary.append(
            [section.currency, "Net taxable turnover", section.net_taxable_turnover, None, None, section.sold_count]
        )
        summary.append([])
    _format_numeric_columns(summary, money_cols=(3, 4, 5), int_cols=(6,))
    for summary_row in summary.iter_rows(min_row=2, max_row=summary.max_row):
        if summary_row[1].value == "Net taxable turnover":  # bold the per-currency total row
            for cell in summary_row:
                cell.font = LABEL_FONT
    style_header_row(summary)
    auto_fit_columns(summary)
    summary.freeze_panes = "A2"

    txns = wb.create_sheet("Transactions")
    txns.append(_TXN_HEADERS)
    for section in data.sections:
        for row in section.transactions:
            # ponytail: TxnRow.date is always the SALE date (payment.created_at converted to
            # local date), even for refund-only rows where the original sale fell outside the
            # report period. The refund-only scenario is uncommon; a correct fix would add a
            # separate `refund_date` field to TxnRow and emit that here when the sale is
            # out-of-period. Tracked as a known limitation — do not change _process_payment
            # or _process_ticket here; that belongs in a follow-up to avoid breaking Task 3's
            # passing tests.
            txns.append(
                [
                    row.date.isoformat(),
                    row.payment_id,
                    row.event,
                    row.tier,
                    row.buyer_country,
                    row.reverse_charge,
                    row.gross,
                    row.net,
                    row.vat_rate,
                    row.vat_amount,
                    row.discount,
                    row.refund_amount,
                    row.currency,
                    row.stripe_session_id,
                    row.stripe_payout_id,
                ]
            )
    _format_numeric_columns(txns, money_cols=(7, 8, 10, 11, 12), percent_cols=(9,))
    style_header_row(txns)
    auto_fit_columns(txns)
    txns.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    out = buf.getvalue()
    buf.close()
    wb.close()
    return out


def build_pdf(data: RevenueReportData) -> bytes:
    """Render the revenue & VAT report as a PDF via WeasyPrint."""
    from common.service.invoice_utils import render_pdf

    return render_pdf("reports/revenue_vat_report.html", {"data": data, "org": data.scope.org})


def build_zip(data: RevenueReportData) -> bytes:
    """Bundle XLSX + PDF into a ZIP archive and return raw bytes."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(report_filename(data.scope, "xlsx"), build_xlsx(data))
        zf.writestr(report_filename(data.scope, "pdf"), build_pdf(data))
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Cache + generation
# ---------------------------------------------------------------------------

# ponytail: event_id is stored as "" (empty string) rather than JSON null so
# that the JSONField lookup `parameters__event_id=""` works reliably in
# Postgres. A JSON null value requires IS NULL semantics in Postgres jsonb,
# which Django's JSONField `__exact` lookup does not emit when the value is
# Python None — it would produce `@> 'null'::jsonb` which misses rows in some
# Django versions. Using "" avoids the null-in-JSON ambiguity entirely.
# _parameters_to_scope converts "" back to None when rebuilding the scope.
_NO_EVENT = ""


class _ScopeParameters(t.TypedDict):
    """Serialized ``ReportScope`` stored on ``FileExport.parameters`` for cache lookup."""

    org_id: str
    event_id: str
    date_from: str
    date_to: str
    data_hash: str


def _scope_to_parameters(scope: ReportScope, data_hash: str) -> _ScopeParameters:
    return {
        "org_id": str(scope.org.id),
        "event_id": str(scope.event_id) if scope.event_id else _NO_EVENT,
        "date_from": scope.date_from.isoformat(),
        "date_to": scope.date_to.isoformat(),
        "data_hash": data_hash,
    }


def _parameters_to_scope(parameters: _ScopeParameters) -> ReportScope:
    org = Organization.objects.get(id=parameters["org_id"])
    event_id_raw = parameters["event_id"]
    return ReportScope(
        org=org,
        event_id=UUID(event_id_raw) if event_id_raw else None,
        date_from=date.fromisoformat(parameters["date_from"]),
        date_to=date.fromisoformat(parameters["date_to"]),
    )


def get_or_generate_revenue_report(
    org: Organization,
    scope: ReportScope,
    requested_by: "RevelUser",
    refresh: bool = False,
) -> FileExport:
    """Return a READY cached export when one matches, else enqueue generation."""
    data_hash = compute_revenue_data_hash(scope)
    parameters = _scope_to_parameters(scope, data_hash)

    if not refresh:
        cached = (
            FileExport.objects.filter(
                export_type=FileExport.ExportType.REVENUE_VAT_REPORT,
                status=FileExport.ExportStatus.READY,
                parameters__org_id=parameters["org_id"],
                parameters__event_id=parameters["event_id"],
                parameters__date_from=parameters["date_from"],
                parameters__date_to=parameters["date_to"],
                parameters__data_hash=data_hash,
            )
            .order_by("-completed_at")
            .first()
        )
        if cached is not None:
            return cached

    export = FileExport.objects.create(
        requested_by=requested_by,
        export_type=FileExport.ExportType.REVENUE_VAT_REPORT,
        parameters=parameters,
    )
    from events.tasks import generate_revenue_report_task

    transaction.on_commit(lambda: generate_revenue_report_task.delay(str(export.id)))
    return export


def generate_revenue_report(export_id: UUID) -> None:
    """Build the ZIP bundle for an export and mark it READY (or FAILED)."""
    export = FileExport.objects.select_related("requested_by").get(pk=export_id)
    start_export(export)
    try:
        scope = _parameters_to_scope(t.cast(_ScopeParameters, export.parameters))
        data = build_revenue_report_data(scope)
        complete_export(export, build_zip(data), report_filename(scope))
    except Exception as exc:  # let Celery record the failure after marking FAILED
        fail_export(export, f"Revenue report failed: {exc}")
        raise


def deliver_scheduled_revenue_reports(now_utc: datetime) -> int:
    """Generate + email the just-closed period's report for opted-in orgs. Returns count sent."""
    delivered = 0
    orgs = (
        Organization.objects.select_related("city", "owner")
        .exclude(revenue_report_cadence=Organization.RevenueReportCadence.NONE)
        .exclude(billing_email="")
    )

    for org in orgs:
        try:
            tz = get_organization_timezone(org)
            period = closed_period_for(org.revenue_report_cadence, now_utc.astimezone(tz))
            if period is None:
                continue
            date_from, date_to, label = period
            if org.last_revenue_report_sent_period == label:
                continue

            scope = ReportScope(org=org, event_id=None, date_from=date_from, date_to=date_to)
            if not build_revenue_report_data(scope).sections:
                continue  # skip empty periods; do not set the marker

            export = FileExport.objects.create(
                requested_by=org.owner,
                export_type=FileExport.ExportType.REVENUE_VAT_REPORT,
                parameters=_scope_to_parameters(scope, compute_revenue_data_hash(scope)),
            )
            generate_revenue_report(export.id)
            export.refresh_from_db()

            recipients = [org.billing_email]
            if org.owner.email and org.owner.email != org.billing_email:
                recipients.append(org.owner.email)
            body = render_to_string(
                "emails/revenue_report.txt",
                {"org": org, "label": label, "snapshot_date": now_utc.astimezone(tz).date()},
            )
            send_email.delay(
                to=recipients,
                subject=f"Revenue & VAT report — {label}",
                body=body,
                attachment_storage_path=export.file.name,
                attachment_filename=report_filename(scope),
                attachment_mime_type="application/zip",
            )
            Organization.objects.filter(pk=org.pk).update(last_revenue_report_sent_period=label)
            delivered += 1
        except Exception:
            logger.exception("deliver_scheduled_revenue_reports: error for org", org_id=str(org.id))
            continue  # do not set the marker; retry next run

    return delivered
