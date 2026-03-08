"""Tests for attendee list Excel export service.

Tests cover:
- Excel content generation with tickets and RSVPs
- Summary sheet statistics (totals, checked-in count)
- Attendees sheet rows for tickets and RSVPs
- Empty attendees case
- Error handling (fail_export on exception)
"""

import typing as t
from io import BytesIO

import pytest
from django.utils import timezone
from openpyxl import load_workbook

from accounts.models import RevelUser
from common.models import FileExport
from conftest import RevelUserFactory
from events.models import (
    Event,
    EventRSVP,
    Organization,
    Ticket,
    TicketTier,
    Venue,
)
from events.service.export.attendee_export import generate_attendee_export

pytestmark = pytest.mark.django_db


# --- Fixtures ---


@pytest.fixture
def export_user(revel_user_factory: RevelUserFactory) -> RevelUser:
    """User who requests the export."""
    return revel_user_factory(username="att_exporter")


@pytest.fixture
def org(export_user: RevelUser) -> Organization:
    """Test organization."""
    return Organization.objects.create(name="Attendee Org", slug="att-org", owner=export_user)


@pytest.fixture
def venue(org: Organization) -> Venue:
    """A venue for the event."""
    return Venue.objects.create(
        organization=org,
        name="Test Venue",
        address="123 Main St",
    )


@pytest.fixture
def att_event(org: Organization, venue: Venue) -> Event:
    """Event for attendee export."""
    return Event.objects.create(
        organization=org,
        name="Attendee Export Event",
        slug="att-export-event",
        event_type=Event.EventType.PUBLIC,
        visibility=Event.Visibility.PUBLIC,
        start=timezone.now(),
        status="open",
        venue=venue,
    )


@pytest.fixture
def free_tier(att_event: Event) -> TicketTier:
    """Free ticket tier."""
    return TicketTier.objects.create(
        event=att_event,
        name="Free Tier",
        payment_method=TicketTier.PaymentMethod.FREE,
    )


@pytest.fixture
def ticket_user(revel_user_factory: RevelUserFactory) -> RevelUser:
    """A user who has a ticket."""
    return revel_user_factory(
        username="ticket_holder",
        email="ticket@example.com",
        first_name="Alice",
        last_name="Ticker",
        pronouns="she/her",
    )


@pytest.fixture
def rsvp_user(revel_user_factory: RevelUserFactory) -> RevelUser:
    """A user who RSVP'd."""
    return revel_user_factory(
        username="rsvp_user",
        email="rsvp@example.com",
        first_name="Bob",
        last_name="Rsvper",
        pronouns="he/him",
    )


@pytest.fixture
def active_ticket(att_event: Event, ticket_user: RevelUser, free_tier: TicketTier) -> Ticket:
    """An active ticket."""
    return Ticket.objects.create(
        event=att_event,
        user=ticket_user,
        tier=free_tier,
        status=Ticket.TicketStatus.ACTIVE,
        guest_name="Alice Ticker",
    )


@pytest.fixture
def checked_in_ticket(
    att_event: Event,
    revel_user_factory: RevelUserFactory,
    free_tier: TicketTier,
) -> Ticket:
    """A checked-in ticket."""
    user = revel_user_factory(
        username="checkedin", email="checkedin@example.com", first_name="Carol", last_name="Check"
    )
    ticket = Ticket.objects.create(
        event=att_event,
        user=user,
        tier=free_tier,
        status=Ticket.TicketStatus.CHECKED_IN,
        guest_name="Carol Check",
    )
    # Manually set checked_in_at since it's editable=False
    Ticket.objects.filter(pk=ticket.pk).update(checked_in_at=timezone.now())
    ticket.refresh_from_db()
    return ticket


@pytest.fixture
def rsvp_yes(att_event: Event, rsvp_user: RevelUser) -> EventRSVP:
    """A YES RSVP."""
    return EventRSVP.objects.create(
        event=att_event,
        user=rsvp_user,
        status=EventRSVP.RsvpStatus.YES,
    )


def _create_attendee_export(export_user: RevelUser, event: Event) -> FileExport:
    """Helper to create a FileExport for attendee list."""
    return FileExport.objects.create(
        requested_by=export_user,
        export_type=FileExport.ExportType.ATTENDEE_LIST,
        parameters={"event_id": str(event.id)},
    )


def _load_workbook_from_export(export: FileExport) -> t.Any:
    """Load the openpyxl workbook from a completed export."""
    export.refresh_from_db()
    content = export.file.read()
    return load_workbook(BytesIO(content))


# --- Summary Sheet Tests ---


class TestAttendeeExportSummary:
    """Tests for the Summary sheet content."""

    def test_summary_sheet_event_info(
        self,
        export_user: RevelUser,
        att_event: Event,
        active_ticket: Ticket,
        rsvp_yes: EventRSVP,
    ) -> None:
        """Summary sheet should contain event metadata."""
        export = _create_attendee_export(export_user, att_event)

        generate_attendee_export(export.id)

        wb = _load_workbook_from_export(export)
        ws_summary = wb["Summary"]
        summary = {row[0]: row[1] for row in ws_summary.iter_rows(values_only=True) if row[0]}

        assert summary["Event"] == "Attendee Export Event"
        assert summary["Venue"] == "Test Venue"
        assert summary["Organization"] == "Attendee Org"

    def test_summary_sheet_counts(
        self,
        export_user: RevelUser,
        att_event: Event,
        active_ticket: Ticket,
        checked_in_ticket: Ticket,
        rsvp_yes: EventRSVP,
    ) -> None:
        """Summary sheet should have correct attendee counts."""
        export = _create_attendee_export(export_user, att_event)

        generate_attendee_export(export.id)

        wb = _load_workbook_from_export(export)
        ws_summary = wb["Summary"]
        summary = {row[0]: row[1] for row in ws_summary.iter_rows(values_only=True) if row[0]}

        assert summary["Tickets"] == 2
        assert summary["RSVPs"] == 1
        assert summary["Total attendees"] == 3
        assert summary["Checked in"] == 1

        # Pronoun distribution
        assert summary["Total with pronouns"] == 3
        assert summary["Total without pronouns"] == 0


# --- Attendees Sheet Tests ---


class TestAttendeeExportAttendees:
    """Tests for the Attendees sheet content."""

    def test_attendees_sheet_headers(
        self,
        export_user: RevelUser,
        att_event: Event,
        active_ticket: Ticket,
    ) -> None:
        """Attendees sheet should have proper headers."""
        export = _create_attendee_export(export_user, att_event)

        generate_attendee_export(export.id)

        wb = _load_workbook_from_export(export)
        ws_att = wb["Attendees"]
        headers = [cell.value for cell in next(ws_att.iter_rows(min_row=1, max_row=1))]

        expected_headers = [
            "Name",
            "Email",
            "Pronouns",
            "Type",
            "RSVP Status",
            "Ticket Tier",
            "Ticket Status",
            "Checked In",
            "Checked In At",
            "Guest Name",
            "Seat",
            "Payment",
        ]
        assert headers == expected_headers

    def test_ticket_row_content(
        self,
        export_user: RevelUser,
        att_event: Event,
        active_ticket: Ticket,
        ticket_user: RevelUser,
    ) -> None:
        """Ticket rows should contain user info, tier, status, and attendance type."""
        export = _create_attendee_export(export_user, att_event)

        generate_attendee_export(export.id)

        wb = _load_workbook_from_export(export)
        ws_att = wb["Attendees"]
        rows = list(ws_att.iter_rows(min_row=2, values_only=True))

        assert len(rows) == 1
        row = rows[0]

        # Name, email, pronouns
        assert "Alice" in str(row[0])  # Name
        assert row[1] == "ticket@example.com"  # Email
        assert row[2] == "she/her"  # Pronouns
        assert row[3] == "Ticket"  # Type
        assert row[5] == "Free Tier"  # Ticket Tier
        assert row[6] == "Active"  # Ticket Status
        assert row[7] == "No"  # Checked In
        assert row[9] == "Alice Ticker"  # Guest Name

    def test_checked_in_ticket_row(
        self,
        export_user: RevelUser,
        att_event: Event,
        checked_in_ticket: Ticket,
    ) -> None:
        """Checked-in tickets should show correct check-in status and timestamp."""
        export = _create_attendee_export(export_user, att_event)

        generate_attendee_export(export.id)

        wb = _load_workbook_from_export(export)
        ws_att = wb["Attendees"]
        rows = list(ws_att.iter_rows(min_row=2, values_only=True))

        assert len(rows) == 1
        row = rows[0]
        assert row[6] == "Checked In"  # Ticket Status
        assert row[7] == "Yes"  # Checked In
        assert row[8] != ""  # Checked In At should have a value

    def test_rsvp_row_content(
        self,
        export_user: RevelUser,
        att_event: Event,
        rsvp_yes: EventRSVP,
        rsvp_user: RevelUser,
    ) -> None:
        """RSVP rows should show user info and attendance type 'rsvp'."""
        export = _create_attendee_export(export_user, att_event)

        generate_attendee_export(export.id)

        wb = _load_workbook_from_export(export)
        ws_att = wb["Attendees"]
        rows = list(ws_att.iter_rows(min_row=2, values_only=True))

        assert len(rows) == 1
        row = rows[0]
        assert "Bob" in str(row[0])  # Name
        assert row[1] == "rsvp@example.com"  # Email
        assert row[2] == "he/him"  # Pronouns
        assert row[3] == "RSVP"  # Type
        assert row[4] == "Yes"  # RSVP Status
        # The remaining fields should be empty/None for RSVPs
        assert not row[5]  # Ticket Tier
        assert not row[6]  # Ticket Status

    def test_cancelled_tickets_excluded(
        self,
        export_user: RevelUser,
        att_event: Event,
        free_tier: TicketTier,
        revel_user_factory: RevelUserFactory,
    ) -> None:
        """Cancelled tickets should be excluded from the attendees list."""
        user = revel_user_factory(username="cancelled_user")
        Ticket.objects.create(
            event=att_event,
            user=user,
            tier=free_tier,
            status=Ticket.TicketStatus.CANCELLED,
            guest_name="Cancelled Guest",
        )

        export = _create_attendee_export(export_user, att_event)

        generate_attendee_export(export.id)

        wb = _load_workbook_from_export(export)
        ws_att = wb["Attendees"]
        rows = list(ws_att.iter_rows(min_row=2, values_only=True))

        assert len(rows) == 0

    def test_mixed_tickets_and_rsvps(
        self,
        export_user: RevelUser,
        att_event: Event,
        active_ticket: Ticket,
        rsvp_yes: EventRSVP,
    ) -> None:
        """Export should include both ticket holders and RSVP users."""
        export = _create_attendee_export(export_user, att_event)

        generate_attendee_export(export.id)

        wb = _load_workbook_from_export(export)
        ws_att = wb["Attendees"]
        rows = list(ws_att.iter_rows(min_row=2, values_only=True))

        assert len(rows) == 2
        attendance_types = [row[3] for row in rows]
        assert "Ticket" in attendance_types
        assert "RSVP" in attendance_types


# --- Empty Attendees Tests ---


class TestAttendeeExportEmpty:
    """Tests for export with no attendees."""

    def test_empty_event_produces_valid_excel(
        self,
        export_user: RevelUser,
        att_event: Event,
    ) -> None:
        """Export for an event with no tickets or RSVPs should still produce valid Excel."""
        export = _create_attendee_export(export_user, att_event)

        generate_attendee_export(export.id)

        export.refresh_from_db()
        assert export.status == FileExport.ExportStatus.READY

        wb = _load_workbook_from_export(export)
        ws_summary = wb["Summary"]
        summary = {row[0]: row[1] for row in ws_summary.iter_rows(values_only=True) if row[0]}
        assert summary["Total attendees"] == 0
        assert summary["Tickets"] == 0
        assert summary["RSVPs"] == 0
        assert summary["Checked in"] == 0
        assert summary["Total with pronouns"] == 0
        assert summary["Total without pronouns"] == 0

    def test_empty_event_has_attendees_sheet(
        self,
        export_user: RevelUser,
        att_event: Event,
    ) -> None:
        """Even with no attendees, the Attendees sheet should exist with headers."""
        export = _create_attendee_export(export_user, att_event)

        generate_attendee_export(export.id)

        wb = _load_workbook_from_export(export)
        ws_att = wb["Attendees"]
        rows = list(ws_att.iter_rows(values_only=True))
        assert len(rows) == 1  # Only header row
        assert rows[0][0] == "Name"


# --- Error Handling Tests ---


class TestAttendeeExportErrorHandling:
    """Tests for error handling during attendee export generation."""

    def test_invalid_event_id_fails_export(self, export_user: RevelUser) -> None:
        """Export with a non-existent event_id should fail gracefully."""
        export = FileExport.objects.create(
            requested_by=export_user,
            export_type=FileExport.ExportType.ATTENDEE_LIST,
            parameters={"event_id": "00000000-0000-0000-0000-000000000099"},
        )

        with pytest.raises(Exception):
            generate_attendee_export(export.id)

        export.refresh_from_db()
        assert export.status == FileExport.ExportStatus.FAILED
        assert export.error_message is not None
        assert export.error_message.startswith("Export failed:")


# --- Event Without Venue Tests ---


class TestAttendeeExportNoVenue:
    """Tests for export when event has no venue."""

    def test_no_venue_shows_na(
        self,
        export_user: RevelUser,
        org: Organization,
    ) -> None:
        """Summary sheet should show 'N/A' for venue when event has no venue."""
        event_no_venue = Event.objects.create(
            organization=org,
            name="No Venue Event",
            slug="no-venue-event",
            event_type=Event.EventType.PUBLIC,
            visibility=Event.Visibility.PUBLIC,
            start=timezone.now(),
            status="open",
        )

        export = _create_attendee_export(export_user, event_no_venue)

        generate_attendee_export(export.id)

        wb = _load_workbook_from_export(export)
        ws_summary = wb["Summary"]
        summary = {row[0]: row[1] for row in ws_summary.iter_rows(values_only=True) if row[0]}
        assert summary["Venue"] == "N/A"
