"""Tests for questionnaire Excel export service.

Tests cover:
- Excel content generation with MC, FT, and FU answers
- Summary sheet statistics (total, unique users, score stats)
- Event filtering (by event_id and event_series_id)
- Empty submissions case
- Error handling (fail_export on exception)
"""

import typing as t
from decimal import Decimal
from io import BytesIO

import pytest
from django.core.files.base import ContentFile
from django.utils import timezone
from openpyxl import load_workbook

from accounts.models import RevelUser
from common.models import FileExport
from conftest import RevelUserFactory
from events.models import (
    Event,
    EventQuestionnaireSubmission,
    EventSeries,
    Organization,
    OrganizationQuestionnaire,
)
from events.service.export.questionnaire_export import generate_questionnaire_export
from questionnaires.models import (
    FileUploadAnswer,
    FileUploadQuestion,
    FreeTextAnswer,
    FreeTextQuestion,
    MultipleChoiceAnswer,
    MultipleChoiceOption,
    MultipleChoiceQuestion,
    Questionnaire,
    QuestionnaireEvaluation,
    QuestionnaireFile,
    QuestionnaireSection,
    QuestionnaireSubmission,
)

pytestmark = pytest.mark.django_db


# --- Fixtures ---


@pytest.fixture
def export_user(revel_user_factory: RevelUserFactory) -> RevelUser:
    """User who requests the export."""
    return revel_user_factory(username="q_exporter")


@pytest.fixture
def submitter(revel_user_factory: RevelUserFactory) -> RevelUser:
    """User who submits questionnaire answers."""
    return revel_user_factory(
        username="submitter",
        email="submitter@example.com",
        first_name="Jane",
        last_name="Smith",
        pronouns="she/her",
    )


@pytest.fixture
def org(export_user: RevelUser) -> Organization:
    """Test organization."""
    return Organization.objects.create(name="Export Org", slug="export-org", owner=export_user)


@pytest.fixture
def q_event_series(org: Organization) -> EventSeries:
    """Event series for filtering tests."""
    return EventSeries.objects.create(organization=org, name="Q Series", slug="q-series")


@pytest.fixture
def q_event(org: Organization, q_event_series: EventSeries) -> Event:
    """Event for filtering tests."""
    return Event.objects.create(
        organization=org,
        name="Q Event",
        slug="q-event",
        event_type=Event.EventType.PUBLIC,
        visibility=Event.Visibility.PUBLIC,
        event_series=q_event_series,
        start=timezone.now(),
        status="open",
    )


@pytest.fixture
def questionnaire() -> Questionnaire:
    """Published questionnaire with a section."""
    return Questionnaire.objects.create(
        name="Export Test Questionnaire",
        status=Questionnaire.QuestionnaireStatus.PUBLISHED,
    )


@pytest.fixture
def section(questionnaire: Questionnaire) -> QuestionnaireSection:
    """A section for the questionnaire."""
    return QuestionnaireSection.objects.create(questionnaire=questionnaire, name="Section 1", order=1)


@pytest.fixture
def mc_question(questionnaire: Questionnaire, section: QuestionnaireSection) -> MultipleChoiceQuestion:
    """Multiple choice question."""
    return MultipleChoiceQuestion.objects.create(
        questionnaire=questionnaire,
        section=section,
        question="Favorite color?",
        order=1,
    )


@pytest.fixture
def mc_options(mc_question: MultipleChoiceQuestion) -> list[MultipleChoiceOption]:
    """Options for the MC question."""
    return [
        MultipleChoiceOption.objects.create(question=mc_question, option="Red", order=1),
        MultipleChoiceOption.objects.create(question=mc_question, option="Blue", order=2),
    ]


@pytest.fixture
def ft_question(questionnaire: Questionnaire, section: QuestionnaireSection) -> FreeTextQuestion:
    """Free text question."""
    return FreeTextQuestion.objects.create(
        questionnaire=questionnaire,
        section=section,
        question="Why do you want to attend?",
        order=2,
    )


@pytest.fixture
def fu_question(questionnaire: Questionnaire, section: QuestionnaireSection) -> FileUploadQuestion:
    """File upload question."""
    return FileUploadQuestion.objects.create(
        questionnaire=questionnaire,
        section=section,
        question="Upload your CV",
        order=3,
    )


@pytest.fixture
def org_questionnaire(org: Organization, questionnaire: Questionnaire) -> OrganizationQuestionnaire:
    """Link questionnaire to organization."""
    return OrganizationQuestionnaire.objects.create(organization=org, questionnaire=questionnaire)


@pytest.fixture
def submission_with_answers(
    submitter: RevelUser,
    questionnaire: Questionnaire,
    mc_question: MultipleChoiceQuestion,
    mc_options: list[MultipleChoiceOption],
    ft_question: FreeTextQuestion,
    fu_question: FileUploadQuestion,
) -> QuestionnaireSubmission:
    """A READY submission with MC, FT, and FU answers."""
    submission = QuestionnaireSubmission.objects.create(
        user=submitter,
        questionnaire=questionnaire,
        status=QuestionnaireSubmission.QuestionnaireSubmissionStatus.READY,
    )

    # MC answer - selecting "Red"
    MultipleChoiceAnswer.objects.create(
        submission=submission,
        question=mc_question,
        option=mc_options[0],
    )

    # FT answer
    FreeTextAnswer.objects.create(
        submission=submission,
        question=ft_question,
        answer="I love music festivals!",
    )

    # FU answer
    fu_answer = FileUploadAnswer.objects.create(
        submission=submission,
        question=fu_question,
    )
    q_file = QuestionnaireFile.objects.create(
        uploader=submitter,
        file=ContentFile(b"fake-pdf", name="cv.pdf"),
        original_filename="my_resume.pdf",
        file_hash="abc123",
        mime_type="application/pdf",
        file_size=8,
    )
    fu_answer.files.add(q_file)

    return submission


@pytest.fixture
def evaluated_submission(
    submission_with_answers: QuestionnaireSubmission,
) -> QuestionnaireEvaluation:
    """An evaluation for the submission."""
    return QuestionnaireEvaluation.objects.create(
        submission=submission_with_answers,
        status=QuestionnaireEvaluation.QuestionnaireEvaluationStatus.APPROVED,
        score=Decimal("85.50"),
        comments="Great answers!",
    )


def _create_export(export_user: RevelUser, questionnaire: Questionnaire, **extra_params: str) -> FileExport:
    """Helper to create a FileExport for questionnaire submissions."""
    params: dict[str, str] = {"questionnaire_id": str(questionnaire.id)}
    params.update(extra_params)
    return FileExport.objects.create(
        requested_by=export_user,
        export_type=FileExport.ExportType.QUESTIONNAIRE_SUBMISSIONS,
        parameters=params,
    )


def _load_workbook_from_export(export: FileExport) -> t.Any:
    """Load the openpyxl workbook from a completed export."""
    export.refresh_from_db()
    content = export.file.read()
    return load_workbook(BytesIO(content))


# --- Excel Content Tests ---


class TestQuestionnaireExportContent:
    """Tests for the Excel content generated by questionnaire export."""

    def test_summary_sheet_statistics(
        self,
        export_user: RevelUser,
        questionnaire: Questionnaire,
        submission_with_answers: QuestionnaireSubmission,
        evaluated_submission: QuestionnaireEvaluation,
    ) -> None:
        """Summary sheet should contain correct aggregate statistics."""
        export = _create_export(export_user, questionnaire)

        generate_questionnaire_export(export.id)

        wb = _load_workbook_from_export(export)
        ws_summary = wb["Summary"]

        # Parse summary into a dict for easy assertions
        summary = {row[0]: row[1] for row in ws_summary.iter_rows(values_only=True) if row[0]}

        assert summary["Total submissions"] == 1
        assert summary["Unique users"] == 1
        assert summary["Approved"] == 1
        assert summary["Rejected"] == 0
        assert summary["Pending review"] == 0
        assert summary["Not evaluated"] == 0
        assert summary["Average score"] == 85.50
        assert summary["Min score"] == 85.50
        assert summary["Max score"] == 85.50

        # Pronoun distribution
        assert summary["Pronoun Distribution"] is None  # header row, value column is empty
        assert summary["Total with pronouns"] == 1
        assert summary["Total without pronouns"] == 0
        assert summary["  she/her"] == 1

    def test_submissions_sheet_headers(
        self,
        export_user: RevelUser,
        questionnaire: Questionnaire,
        submission_with_answers: QuestionnaireSubmission,
        mc_question: MultipleChoiceQuestion,
        ft_question: FreeTextQuestion,
        fu_question: FileUploadQuestion,
    ) -> None:
        """Submissions sheet should have proper headers including question columns."""
        export = _create_export(export_user, questionnaire)

        generate_questionnaire_export(export.id)

        wb = _load_workbook_from_export(export)
        ws_subs = wb["Submissions"]
        headers = [cell.value for cell in next(ws_subs.iter_rows(min_row=1, max_row=1))]

        # Fixed headers
        assert "Email" in headers
        assert "Name" in headers
        assert "Submitted At" in headers
        assert "Status" in headers
        assert "Score" in headers
        assert "Evaluator Comments" in headers
        assert "Source Event" in headers

        # Question headers
        assert any("[MC]" in h for h in headers if h)
        assert any("[FT]" in h for h in headers if h)
        assert any("[FU]" in h for h in headers if h)

    def test_submissions_sheet_mc_answer(
        self,
        export_user: RevelUser,
        questionnaire: Questionnaire,
        submission_with_answers: QuestionnaireSubmission,
        mc_question: MultipleChoiceQuestion,
        mc_options: list[MultipleChoiceOption],
        ft_question: FreeTextQuestion,
        fu_question: FileUploadQuestion,
        evaluated_submission: QuestionnaireEvaluation,
    ) -> None:
        """Submissions sheet should contain the selected MC option text."""
        export = _create_export(export_user, questionnaire)

        generate_questionnaire_export(export.id)

        wb = _load_workbook_from_export(export)
        ws_subs = wb["Submissions"]
        rows = list(ws_subs.iter_rows(min_row=2, values_only=True))

        assert len(rows) == 1
        row = rows[0]
        # The MC answer column should contain "Red"
        assert "Red" in row

    def test_submissions_sheet_ft_answer(
        self,
        export_user: RevelUser,
        questionnaire: Questionnaire,
        submission_with_answers: QuestionnaireSubmission,
        mc_question: MultipleChoiceQuestion,
        ft_question: FreeTextQuestion,
        fu_question: FileUploadQuestion,
    ) -> None:
        """Submissions sheet should contain the free text answer."""
        export = _create_export(export_user, questionnaire)

        generate_questionnaire_export(export.id)

        wb = _load_workbook_from_export(export)
        ws_subs = wb["Submissions"]
        rows = list(ws_subs.iter_rows(min_row=2, values_only=True))

        assert len(rows) == 1
        row = rows[0]
        assert "I love music festivals!" in row

    def test_submissions_sheet_fu_answer(
        self,
        export_user: RevelUser,
        questionnaire: Questionnaire,
        submission_with_answers: QuestionnaireSubmission,
        mc_question: MultipleChoiceQuestion,
        ft_question: FreeTextQuestion,
        fu_question: FileUploadQuestion,
    ) -> None:
        """Submissions sheet should contain file upload original filenames."""
        export = _create_export(export_user, questionnaire)

        generate_questionnaire_export(export.id)

        wb = _load_workbook_from_export(export)
        ws_subs = wb["Submissions"]
        rows = list(ws_subs.iter_rows(min_row=2, values_only=True))

        assert len(rows) == 1
        row = rows[0]
        assert "my_resume.pdf" in row

    def test_submissions_sheet_user_info(
        self,
        export_user: RevelUser,
        questionnaire: Questionnaire,
        submission_with_answers: QuestionnaireSubmission,
        submitter: RevelUser,
        mc_question: MultipleChoiceQuestion,
        ft_question: FreeTextQuestion,
        fu_question: FileUploadQuestion,
    ) -> None:
        """Submissions sheet should include user email and full name."""
        export = _create_export(export_user, questionnaire)

        generate_questionnaire_export(export.id)

        wb = _load_workbook_from_export(export)
        ws_subs = wb["Submissions"]
        rows = list(ws_subs.iter_rows(min_row=2, values_only=True))

        row = rows[0]
        assert row[0] == "submitter@example.com"  # User Email
        assert "Jane" in str(row[1])  # User Name contains first name

    def test_submissions_sheet_eval_info(
        self,
        export_user: RevelUser,
        questionnaire: Questionnaire,
        submission_with_answers: QuestionnaireSubmission,
        evaluated_submission: QuestionnaireEvaluation,
        mc_question: MultipleChoiceQuestion,
        ft_question: FreeTextQuestion,
        fu_question: FileUploadQuestion,
    ) -> None:
        """Submissions sheet should show evaluation status, score, and comments."""
        export = _create_export(export_user, questionnaire)

        generate_questionnaire_export(export.id)

        wb = _load_workbook_from_export(export)
        ws_subs = wb["Submissions"]
        rows = list(ws_subs.iter_rows(min_row=2, values_only=True))

        row = rows[0]
        # Eval Status is at index 3, Score at index 4, Comments at index 5
        assert row[3] == "Approved"
        assert row[4] == 85.50
        assert row[5] == "Great answers!"


# --- Event Filtering Tests ---


class TestQuestionnaireExportFiltering:
    """Tests for filtering exports by event_id or event_series_id."""

    def test_filter_by_event_id(
        self,
        export_user: RevelUser,
        questionnaire: Questionnaire,
        q_event: Event,
        org_questionnaire: OrganizationQuestionnaire,
        submitter: RevelUser,
        revel_user_factory: RevelUserFactory,
    ) -> None:
        """Export filtered by event_id should only include submissions linked to that event."""
        # Submission linked to the event
        linked_sub = QuestionnaireSubmission.objects.create(
            user=submitter,
            questionnaire=questionnaire,
            status=QuestionnaireSubmission.QuestionnaireSubmissionStatus.READY,
        )
        EventQuestionnaireSubmission.objects.create(
            event=q_event,
            user=submitter,
            questionnaire=questionnaire,
            submission=linked_sub,
            questionnaire_type=org_questionnaire.questionnaire_type,
        )

        # Unlinked submission (should be excluded)
        other_user = revel_user_factory(username="other_sub")
        QuestionnaireSubmission.objects.create(
            user=other_user,
            questionnaire=questionnaire,
            status=QuestionnaireSubmission.QuestionnaireSubmissionStatus.READY,
        )

        export = _create_export(export_user, questionnaire, event_id=str(q_event.id))

        generate_questionnaire_export(export.id)

        wb = _load_workbook_from_export(export)
        ws_summary = wb["Summary"]
        summary = {row[0]: row[1] for row in ws_summary.iter_rows(values_only=True) if row[0]}
        assert summary["Total submissions"] == 1

    def test_filter_by_event_series_id(
        self,
        export_user: RevelUser,
        questionnaire: Questionnaire,
        q_event: Event,
        q_event_series: EventSeries,
        org_questionnaire: OrganizationQuestionnaire,
        submitter: RevelUser,
        revel_user_factory: RevelUserFactory,
    ) -> None:
        """Export filtered by event_series_id should include submissions linked to events in that series."""
        # Submission linked to an event in the series
        linked_sub = QuestionnaireSubmission.objects.create(
            user=submitter,
            questionnaire=questionnaire,
            status=QuestionnaireSubmission.QuestionnaireSubmissionStatus.READY,
        )
        EventQuestionnaireSubmission.objects.create(
            event=q_event,
            user=submitter,
            questionnaire=questionnaire,
            submission=linked_sub,
            questionnaire_type=org_questionnaire.questionnaire_type,
        )

        # Unlinked submission
        other_user = revel_user_factory(username="series_other")
        QuestionnaireSubmission.objects.create(
            user=other_user,
            questionnaire=questionnaire,
            status=QuestionnaireSubmission.QuestionnaireSubmissionStatus.READY,
        )

        export = _create_export(export_user, questionnaire, event_series_id=str(q_event_series.id))

        generate_questionnaire_export(export.id)

        wb = _load_workbook_from_export(export)
        ws_summary = wb["Summary"]
        summary = {row[0]: row[1] for row in ws_summary.iter_rows(values_only=True) if row[0]}
        assert summary["Total submissions"] == 1


# --- Empty Submissions Tests ---


class TestQuestionnaireExportEmpty:
    """Tests for export with no submissions."""

    def test_empty_submissions_produces_valid_excel(
        self,
        export_user: RevelUser,
        questionnaire: Questionnaire,
    ) -> None:
        """Export with zero submissions should still produce a valid Excel file."""
        export = _create_export(export_user, questionnaire)

        generate_questionnaire_export(export.id)

        export.refresh_from_db()
        assert export.status == FileExport.ExportStatus.READY

        wb = _load_workbook_from_export(export)
        ws_summary = wb["Summary"]
        summary = {row[0]: row[1] for row in ws_summary.iter_rows(values_only=True) if row[0]}
        assert summary["Total submissions"] == 0
        assert summary["Unique users"] == 0
        assert summary["Average score"] == "N/A"
        assert summary["Total with pronouns"] == 0
        assert summary["Total without pronouns"] == 0

    def test_empty_submissions_has_submissions_sheet(
        self,
        export_user: RevelUser,
        questionnaire: Questionnaire,
    ) -> None:
        """Export with zero submissions should still have a Submissions sheet with headers."""
        export = _create_export(export_user, questionnaire)

        generate_questionnaire_export(export.id)

        wb = _load_workbook_from_export(export)
        ws_subs = wb["Submissions"]
        rows = list(ws_subs.iter_rows(values_only=True))
        # Should have at least the header row
        assert len(rows) >= 1
        assert rows[0][0] == "Email"


# --- Error Handling Tests ---


class TestQuestionnaireExportErrorHandling:
    """Tests for error handling during export generation."""

    def test_invalid_questionnaire_id_fails_export(self, export_user: RevelUser) -> None:
        """Export with an invalid questionnaire_id should fail gracefully."""
        export = FileExport.objects.create(
            requested_by=export_user,
            export_type=FileExport.ExportType.QUESTIONNAIRE_SUBMISSIONS,
            parameters={"questionnaire_id": "not-a-uuid"},
        )

        # The generate function catches exceptions and calls fail_export,
        # then re-raises. So we expect the exception to propagate.
        with pytest.raises(Exception):
            generate_questionnaire_export(export.id)

        export.refresh_from_db()
        assert export.status == FileExport.ExportStatus.FAILED
        assert export.error_message is not None
        assert export.error_message.startswith("Export failed:")


# --- Multiple Submissions Tests ---


class TestQuestionnaireExportMultipleSubmissions:
    """Tests for exports with multiple submissions and varied evaluation states."""

    def test_multiple_submissions_with_mixed_evaluations(
        self,
        export_user: RevelUser,
        questionnaire: Questionnaire,
        revel_user_factory: RevelUserFactory,
    ) -> None:
        """Export should correctly aggregate multiple submissions with different eval states."""
        # Create 3 submissions: approved, rejected, no evaluation
        users = [revel_user_factory(username=f"multi_{i}") for i in range(3)]

        sub1 = QuestionnaireSubmission.objects.create(
            user=users[0],
            questionnaire=questionnaire,
            status=QuestionnaireSubmission.QuestionnaireSubmissionStatus.READY,
        )
        QuestionnaireEvaluation.objects.create(
            submission=sub1,
            status=QuestionnaireEvaluation.QuestionnaireEvaluationStatus.APPROVED,
            score=Decimal("90.00"),
        )

        sub2 = QuestionnaireSubmission.objects.create(
            user=users[1],
            questionnaire=questionnaire,
            status=QuestionnaireSubmission.QuestionnaireSubmissionStatus.READY,
        )
        QuestionnaireEvaluation.objects.create(
            submission=sub2,
            status=QuestionnaireEvaluation.QuestionnaireEvaluationStatus.REJECTED,
            score=Decimal("30.00"),
        )

        QuestionnaireSubmission.objects.create(
            user=users[2],
            questionnaire=questionnaire,
            status=QuestionnaireSubmission.QuestionnaireSubmissionStatus.READY,
        )

        export = _create_export(export_user, questionnaire)

        generate_questionnaire_export(export.id)

        wb = _load_workbook_from_export(export)
        ws_summary = wb["Summary"]
        summary = {row[0]: row[1] for row in ws_summary.iter_rows(values_only=True) if row[0]}

        assert summary["Total submissions"] == 3
        assert summary["Unique users"] == 3
        assert summary["Approved"] == 1
        assert summary["Rejected"] == 1
        assert summary["Not evaluated"] == 1
        assert summary["Average score"] == 60.00  # (90 + 30) / 2
        # Factory users get random pronouns by default
        assert summary["Total with pronouns"] == 3
        assert summary["Total without pronouns"] == 0

        ws_subs = wb["Submissions"]
        data_rows = list(ws_subs.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 3
