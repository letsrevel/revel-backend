"""Bundle builder tests for the revenue & VAT report (#551)."""

import datetime as dt
import io
import typing as t
import zipfile
from decimal import Decimal

import pytest
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl import load_workbook

from accounts.models import RevelUser
from events.models import (
    Event,
    MembershipPayment,
    MembershipSubscription,
    MembershipSubscriptionPlan,
    MembershipTier,
    Organization,
    Payment,
    Ticket,
    TicketTier,
)
from events.service import revenue_report_service as svc


@pytest.fixture
def report_data(db: t.Any) -> svc.RevenueReportData:
    user = RevelUser.objects.create_user(username="o", email="o@example.com", password="x")
    org = Organization.objects.create(
        name="Org",
        slug="org",
        owner=user,
        vat_rate=Decimal("20.00"),
        vat_country_code="AT",
        billing_name="Org GmbH",
        vat_id="ATU12345678",
    )
    now = dt.datetime(2026, 6, 1, 12, 0, tzinfo=dt.timezone.utc)
    event = Event.objects.create(
        organization=org,
        name="E",
        slug="e",
        start=now,
        end=now + dt.timedelta(hours=2),
    )
    tier = TicketTier.objects.create(
        event=event,
        name="GA",
        price=Decimal("120.00"),
        currency="EUR",
        payment_method=TicketTier.PaymentMethod.ONLINE,
    )
    ticket = Ticket.objects.create(
        event=event, tier=tier, user=user, status=Ticket.TicketStatus.ACTIVE, guest_name="Alice"
    )
    Payment.objects.create(
        ticket=ticket,
        user=user,
        status=Payment.PaymentStatus.SUCCEEDED,
        amount=Decimal("120.00"),
        currency="EUR",
        net_amount=Decimal("100.00"),
        vat_amount=Decimal("20.00"),
        vat_rate=Decimal("20.00"),
        platform_fee=Decimal("0.00"),
        stripe_session_id="cs_test_bundle",
    )
    # Wide window so the now-stamped sale always falls in-period (year-agnostic).
    scope = svc.ReportScope(org=org, event_id=None, date_from=dt.date(2000, 1, 1), date_to=dt.date(2100, 1, 1))
    return svc.build_revenue_report_data(scope)


@pytest.mark.django_db
def test_xlsx_has_summary_and_transactions_sheets(report_data: svc.RevenueReportData) -> None:
    wb = load_workbook(io.BytesIO(svc.build_xlsx(report_data)))
    assert wb.sheetnames == ["Summary", "Transactions", "Membership payments"]
    headers = [c.value for c in wb["Transactions"][1]]
    assert "payment_id" in headers and "vat_rate" in headers and "stripe_payout_id" in headers


@pytest.mark.django_db
def test_xlsx_is_styled_and_number_formatted(report_data: svc.RevenueReportData) -> None:
    """Money cells carry a thousands-separator format and the VAT label is readable (#554)."""
    wb = load_workbook(io.BytesIO(svc.build_xlsx(report_data)))
    summary = wb["Summary"]
    # Row 2 is the single 20% bucket: B=label, C=Net, E=Gross, F=Tickets.
    assert summary["B2"].value == "20%"
    assert summary["C2"].number_format == "#,##0.00"
    assert summary["F2"].number_format == "#,##0"
    assert summary.freeze_panes == "A2"
    # Header row is styled (bold white-on-blue fill applied by style_header_row).
    assert summary["A1"].font.bold is True

    txns = wb["Transactions"]
    assert txns["G2"].number_format == "#,##0.00"  # gross
    assert txns["I2"].number_format == '0.##"%"'  # vat_rate


@pytest.fixture
def membership_report_data(db: t.Any) -> svc.RevenueReportData:
    """An org whose only money is a membership payment (one refunded, one clean)."""
    owner = RevelUser.objects.create_user(username="mo", email="mo@example.com", password="x")
    member = RevelUser.objects.create_user(
        username="mm", email="member@example.com", password="x", first_name="Mia", last_name="Member"
    )
    org = Organization.objects.create(name="MOrg", slug="morg", owner=owner, vat_country_code="AT")
    tier = MembershipTier.objects.get(organization=org, name="General membership")
    plan = MembershipSubscriptionPlan.objects.create(
        tier=tier, name="Gold", price=Decimal("30.00"), currency="EUR", period_unit="month"
    )
    now = timezone.now()
    sub = MembershipSubscription.objects.create(
        organization=org,
        user=member,
        plan=plan,
        status=MembershipSubscription.SubscriptionStatus.ACTIVE,
        current_period_start=now,
        current_period_end=now + dt.timedelta(days=30),
    )
    MembershipPayment.objects.create(
        subscription=sub,
        amount=Decimal("30.00"),
        currency="EUR",
        status=MembershipPayment.PaymentStatus.SUCCEEDED,
        period_start=now,
        period_end=now + dt.timedelta(days=30),
        platform_fee=Decimal("1.00"),
        refund_amount=Decimal("5.00"),
        refunded_at=now,
        stripe_invoice_id="in_test_sheet",
        stripe_payment_intent_id="pi_test_sheet",
    )
    scope = svc.ReportScope(org=org, event_id=None, date_from=dt.date(2000, 1, 1), date_to=dt.date(2100, 1, 1))
    return svc.build_revenue_report_data(scope)


@pytest.mark.django_db
def test_membership_sheet_carries_the_reconciliation_columns(
    membership_report_data: svc.RevenueReportData,
) -> None:
    """The membership sheet lists the payment id, member, plan, amount, refund and Stripe ids."""
    wb = load_workbook(io.BytesIO(svc.build_xlsx(membership_report_data)))
    sheet = wb["Membership payments"]
    assert [c.value for c in sheet[1]] == [
        "date",
        "payment_id",
        "member_email",
        "member_name",
        "plan",
        "amount",
        "currency",
        "status",
        "refund_amount",
        "stripe_invoice_id",
        "stripe_payment_intent_id",
    ]
    row = [c.value for c in sheet[2]]
    assert row[2:] == [
        "member@example.com",
        "Mia Member",
        "Gold",
        30.0,
        "EUR",
        "succeeded",
        5.0,
        "in_test_sheet",
        "pi_test_sheet",
    ]
    assert sheet["F2"].number_format == "#,##0.00"  # amount


@pytest.mark.django_db
def test_membership_sheet_labels_revenue_as_gross_and_out_of_scope(
    membership_report_data: svc.RevenueReportData,
) -> None:
    """The sheet says membership money is gross and excluded from taxable turnover."""
    wb = load_workbook(io.BytesIO(svc.build_xlsx(membership_report_data)))
    sheet = wb["Membership payments"]
    notes = [
        str(row[0].value)
        for row in sheet.iter_rows(min_row=2, max_col=1)
        if row[0].value and "membership revenue is reported gross" in str(row[0].value)
    ]
    assert len(notes) == 1, "membership VAT note missing from the sheet"
    assert "no VAT treatment is applied" in notes[0]
    assert "excluded from the Net taxable turnover figures" in notes[0]
    assert "tax advisor" in notes[0]
    # The note sits below the data block: headers and the first payment row are untouched,
    # so machine parsers keyed off row 1 / row 2 keep working.
    assert sheet["A1"].value == "date"
    assert sheet["C2"].value == "member@example.com"
    assert sheet["A2"].value == membership_report_data.membership_payments[0].date.isoformat()


@pytest.mark.django_db
def test_membership_only_org_still_produces_a_report(membership_report_data: svc.RevenueReportData) -> None:
    """No tickets at all: the ticket sections are empty but the ledger is not."""
    assert membership_report_data.sections == []
    assert len(membership_report_data.membership_payments) == 1
    assert [m.gross for m in membership_report_data.memberships] == [Decimal("30.00")]


_MAY = dt.datetime(2026, 5, 15, 12, 0, tzinfo=dt.timezone.utc)
_JUNE_2ND = dt.datetime(2026, 6, 2, 12, 0, tzinfo=dt.timezone.utc)
_JUNE_15TH = dt.datetime(2026, 6, 15, 12, 0, tzinfo=dt.timezone.utc)
_JUNE_SCOPE = (dt.date(2026, 6, 1), dt.date(2026, 6, 30))


def _org_with_membership_payment(
    slug: str, *, occurred_at: dt.datetime, refunded_at: dt.datetime
) -> tuple[Organization, MembershipPayment]:
    """An org whose only money is one membership payment, sold and refunded at fixed instants."""
    owner = RevelUser.objects.create_user(username=f"{slug}-o", email=f"{slug}-o@example.com", password="x")
    member = RevelUser.objects.create_user(username=f"{slug}-m", email=f"{slug}-m@example.com", password="x")
    org = Organization.objects.create(name=slug, slug=slug, owner=owner, vat_country_code="AT")
    tier = MembershipTier.objects.get(organization=org, name="General membership")
    plan = MembershipSubscriptionPlan.objects.create(
        tier=tier, name="Gold", price=Decimal("30.00"), currency="EUR", period_unit="month"
    )
    sub = MembershipSubscription.objects.create(
        organization=org,
        user=member,
        plan=plan,
        status=MembershipSubscription.SubscriptionStatus.ACTIVE,
        current_period_start=occurred_at,
        current_period_end=occurred_at + dt.timedelta(days=30),
    )
    payment = MembershipPayment.objects.create(
        subscription=sub,
        amount=Decimal("30.00"),
        currency="EUR",
        status=MembershipPayment.PaymentStatus.REFUNDED,
        period_start=occurred_at,
        period_end=occurred_at + dt.timedelta(days=30),
        occurred_at=occurred_at,
        refund_amount=Decimal("30.00"),
        refunded_at=refunded_at,
    )
    return org, payment


@pytest.mark.django_db
def test_refund_only_membership_row_carries_the_refund_date() -> None:
    """Sold in May, refunded in June: June's ledger line must not be stamped with the May sale date."""
    org, payment = _org_with_membership_payment("refund-only", occurred_at=_MAY, refunded_at=_JUNE_15TH)
    scope = svc.ReportScope(org=org, event_id=None, date_from=_JUNE_SCOPE[0], date_to=_JUNE_SCOPE[1])

    data = svc.build_revenue_report_data(scope)

    (row,) = data.membership_payments
    assert row.date == dt.date(2026, 6, 15)
    assert row.refund_amount == Decimal("30.00")
    assert row.payment_id == str(payment.id)
    # Only the refund is in scope, so the sale contributes nothing to the period's gross.
    assert [m.gross for m in data.memberships] == [Decimal("0.00")]
    # And the sheet the organizer actually opens carries that same in-period date.
    sheet = load_workbook(io.BytesIO(svc.build_xlsx(data)))["Membership payments"]
    assert sheet["A2"].value == "2026-06-15"


@pytest.mark.django_db
def test_membership_row_sold_and_refunded_in_period_keeps_the_sale_date() -> None:
    """Both legs in scope: the line keeps the sale date, matching where the gross is attributed."""
    org, _ = _org_with_membership_payment("same-period", occurred_at=_JUNE_2ND, refunded_at=_JUNE_15TH)
    scope = svc.ReportScope(org=org, event_id=None, date_from=_JUNE_SCOPE[0], date_to=_JUNE_SCOPE[1])

    data = svc.build_revenue_report_data(scope)

    (row,) = data.membership_payments
    assert row.date == dt.date(2026, 6, 2)
    assert [m.gross for m in data.memberships] == [Decimal("30.00")]


@pytest.mark.django_db
def test_offline_membership_payment_is_still_identifiable_on_the_sheet() -> None:
    """No Stripe ids at all: payment_id is the only join key back to the API ledger."""
    org, payment = _org_with_membership_payment("offline-id", occurred_at=_JUNE_2ND, refunded_at=_JUNE_15TH)
    scope = svc.ReportScope(org=org, event_id=None, date_from=_JUNE_SCOPE[0], date_to=_JUNE_SCOPE[1])

    sheet = load_workbook(io.BytesIO(svc.build_xlsx(svc.build_revenue_report_data(scope))))["Membership payments"]

    assert sheet["J2"].value in (None, "")  # stripe_invoice_id
    assert sheet["K2"].value in (None, "")  # stripe_payment_intent_id
    assert sheet["B2"].value == str(payment.id)


@pytest.mark.django_db
def test_pdf_shows_membership_summary_instead_of_no_revenue(
    membership_report_data: svc.RevenueReportData,
) -> None:
    """A membership-only org must not be told "No revenue in this period" (the mailed PDF)."""
    html = render_to_string(
        "reports/revenue_vat_report.html",
        {"data": membership_report_data, "org": membership_report_data.scope.org},
    )
    assert "Memberships" in html
    assert "No revenue in this period" not in html
    assert "30.00" in html  # gross
    assert "5.00" in html  # refunded


@pytest.mark.django_db
def test_pdf_labels_membership_revenue_as_gross_and_out_of_scope(
    membership_report_data: svc.RevenueReportData,
) -> None:
    """The membership table carries the gross / not-taxable-turnover caveat."""
    html = render_to_string(
        "reports/revenue_vat_report.html",
        {"data": membership_report_data, "org": membership_report_data.scope.org},
    )
    assert "Membership revenue is reported gross" in html
    assert "no VAT treatment is applied" in html
    assert "excluded from the net taxable turnover figures" in html
    assert "Consult your tax advisor." in html


@pytest.mark.django_db
def test_pdf_omits_the_membership_table_without_memberships(report_data: svc.RevenueReportData) -> None:
    """Ticket-only orgs see the report exactly as before: no membership table, no membership caveat."""
    html = render_to_string("reports/revenue_vat_report.html", {"data": report_data, "org": report_data.scope.org})
    assert "Memberships" not in html
    assert "No revenue in this period" not in html
    assert "Membership revenue is reported gross" not in html


@pytest.mark.django_db
def test_event_scoped_report_carries_no_membership_rows(membership_report_data: svc.RevenueReportData) -> None:
    """Membership money is org-level: an event-scoped report must not claim it."""
    org = membership_report_data.scope.org
    now = timezone.now()
    event = Event.objects.create(
        organization=org, name="Scoped", slug="scoped", start=now, end=now + dt.timedelta(hours=1)
    )
    scope = svc.ReportScope(org=org, event_id=event.id, date_from=dt.date(2000, 1, 1), date_to=dt.date(2100, 1, 1))

    data = svc.build_revenue_report_data(scope)

    assert data.membership_payments == []
    assert data.memberships == []


@pytest.mark.django_db
def test_pdf_is_nonempty_pdf(report_data: svc.RevenueReportData) -> None:
    pdf = svc.build_pdf(report_data)
    assert pdf[:4] == b"%PDF"


@pytest.mark.django_db
def test_zip_contains_exactly_xlsx_and_pdf(report_data: svc.RevenueReportData) -> None:
    with zipfile.ZipFile(io.BytesIO(svc.build_zip(report_data))) as zf:
        names = sorted(zf.namelist())
    assert len(names) == 2
    assert any(n.endswith(".xlsx") for n in names)
    assert any(n.endswith(".pdf") for n in names)
